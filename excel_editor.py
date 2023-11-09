from openpyxl import load_workbook
import json
import pandas as pd

def get_name_dict():
    with open('gene_names.json') as json_file:
        names_data = json.load(json_file)
        return names_data

def change_gene_names(names_data):
    #load excel file
    print ("Begin load...")
    wb = load_workbook(filename="MAGMA_Gene_Level_Analysis_Results_03_03_23.xlsx")
    print ("Load completed.")
    for sheet in wb:
        print(sheet.title)
        #no cells to change in first sheet
        if sheet.title == "Overview" or sheet["A1"].value == "Name":
            continue
        sheet.insert_cols(1)
        sheet["A1"] = "Name"
        #modify row
        for a,b in sheet.iter_rows(min_row=2, min_col=1, max_col=2):
            name = names_data[str(b.value)]
            # print(name)
            a.value = name
    wb.save("MAGMA_Gene_Level_Analysis_Results_03_03_23_Modified.xlsx")
    print("Done")


def compare_NCBI_builds (names_data):
    df = pd.read_csv("TopTenPercent/disorders/ALZHEIMER_SUMSTATS.txt_step2.genes.out", sep=r"\s+")
    df2 = pd.read_csv('TopTenPercent/disorders/CIGDAY_CUT_SUMSTATS_step2.genes.out', sep=r"\s+")
    df3 = pd.read_csv('TopTenPercent/disorders/ALZHEIMER_SUMSTATS.txt_step2.genes.out', sep=r"\s+")
    # print (set(df['GENE']) - set(names_data.keys()))
    print (len (set(df['GENE'])))
    print (len (set(df2['GENE'])))
    print (len (set(df3['GENE'])))
    # print (len (set(names_data.keys())))
    # print (set([str(gene) for gene in df['GENE']]) - set(names_data.keys()))
    print (len((set([str(gene) for gene in df['GENE']]) | set([str(gene) for gene in df2['GENE']]) | set([str(gene) for gene in df3['GENE']])) - set(names_data.keys())))
    print ('105376335' in names_data.keys())
    print ('5805' in names_data.keys())

    xls = pd.ExcelFile("TopTenPercent/MAGMA_Gene_Level_Analysis_Results_03_03_23_Modified.xlsx")
    df4 = pd.read_excel(xls, "SCZ_2022")
    print (len((set([str(gene) for gene in df['GENE']]) | set([str(gene) for gene in df2['GENE']]) | set([str(gene) for gene in df3['GENE']])) - set([str(gene) for gene in df4['GENE']])))



if __name__ == "__main__":
    names_data = get_name_dict()
    # change_gene_names(names_data)
    compare_NCBI_builds (names_data)
