from openpyxl import load_workbook
import re
from collections import defaultdict
import matplotlib.pyplot as plt
import numpy as np
from scipy.stats import chisquare

def percent_to_float(percent):
    return float(percent.strip('%'))/100

def reorder_dict(dict1, dict2, debugFlag=False):
    list1 = sorted(dict1.items(), key= lambda x:-x[1])
    labels = sorted(dict1.keys(), key= lambda x:-dict1[x])
    if not debugFlag:
        list2 = sorted(dict2.items(), key= lambda x: labels.index(x[0]))
    else:
        list2 = []
        for label in labels:
            if label in dict2.keys():
                list2.append((label, dict2[label]))
    return list1, list2

def chi_square(dict1, dict2, cleaning=False):    
    labels = sorted(dict1.keys(), key= lambda x:-dict1[x])
    if cleaning: # dict1 contains the superset of keys of dict2
        for label in labels:
            if label not in dict2.keys():
                dict2[label] = 0
    list1 = np.array(list(map(lambda x: x[1], sorted(dict1.items(), key= lambda x:-x[1]))))
    list2 = np.array(list(map(lambda x: x[1], sorted(dict2.items(), key= lambda x: labels.index(x[0])))))

    sum1 = np.sum(list1)
    sum2 = np.sum(list2)
    assert sum1 != 0
    assert sum2 != 0

    # want expected frequency to be > 5
    list1 = np.divide(list1, sum1)
    scalar = 5/ np.amin(list1)
    list1 = np.multiply(list1, scalar)
    list2 = np.divide(list2, (sum2/scalar))

    return chisquare(list2, f_exp=list1)

def get_all_dissections (check_against_dataset = False):
    all_dissections = set()
    print ("loading...")
    wb = load_workbook(filename="EDITED_SILETTI_MAGMA_GSA_12_18_22.xlsx")
    print ("loaded")

    for sheet in wb:
        for row in sheet.iter_rows(min_row=2, min_col=18, max_col=18):
            for cell in row:
                dissections = re.split(':|, ', cell.value)
                for i in range (0, len(dissections), 2):
                    all_dissections.add (dissections[i])
    print (all_dissections)
    print (len(all_dissections))
    if check_against_dataset:
        data_set_dissections = set ()
        with open('Siletti_dataset_ROI_values.txt', 'r') as file:
            for line in file:
                line = re.split('"|“', line.strip())
                for dissection in line[1:]:
                    if dissection == "" or dissection == " ":
                        continue
                    data_set_dissections.add(dissection)
        # print (data_set_dissections)
        # print (len(data_set_dissections))
        print (all_dissections - data_set_dissections)
        print (data_set_dissections - all_dissections)
    return all_dissections

def get_cell_breakdown(significance, disorder="SCZ_2022", edit = True):

    wb = load_workbook(filename="EDITED_SILETTI_MAGMA_GSA_12_18_22.xlsx")
    sheet = wb[disorder]

    cuttoff = 0
    break_flag = False
    #find the cuttoff row for significance
    for row in sheet.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            if cell.value > significance:
                break_flag = True
                break
            cuttoff +=1 
        if break_flag == True:
            break
    if cuttoff == 0:
        print(f"There are {cuttoff} significant clusters at the alpha={significance} level! \n")
        return 

    # Column Q
    all_region_three_count = defaultdict(int)
    sig_region_three_count = defaultdict(int)
    all_region_three_weighted = defaultdict(float)
    sig_region_three_weighted = defaultdict(float)
    all_region_top_weighted = defaultdict(float)
    sig_region_top_weighted = defaultdict(float)
    weighted_region_three_weighted = defaultdict(float)
    log_weighted_region_three_weighted = defaultdict(float)
    for row in sheet.iter_rows(min_row=2, min_col=17, max_col=17):
        for cell in row:
            data = cell.value 
            region_values = re.split(':|, ', data)
            p_value = sheet[f"G{cell.row }"].value
            # top one
            all_region_top_weighted[region_values[0]] += percent_to_float(region_values[1])
            if cell.row <= cuttoff+1:
                sig_region_top_weighted[region_values[0]] += percent_to_float(region_values[1])
            # top three
            for i in range (0, len(region_values), 2):
                # all regions
                all_region_three_count[region_values[i]] += 1
                all_region_three_weighted[region_values[i]] += percent_to_float(region_values[i+1])
                # sig regions
                if cell.row <= cuttoff+1:
                    sig_region_three_count[region_values[i]] += 1
                    sig_region_three_weighted[region_values[i]] += percent_to_float(region_values[i+1])
                # proability weighted for regions
                weighted_region_three_weighted[region_values[i]] += percent_to_float(region_values[i+1])/p_value
                log_weighted_region_three_weighted[region_values[i]] -= percent_to_float(region_values[i+1]) * np.log(p_value)
    
    # Column R
    all_dissections_three_count = defaultdict(int)
    sig_dissections_three_count = defaultdict(int)
    all_dissections_three_weighted = defaultdict(float)
    sig_dissections_three_weighted = defaultdict(float)
    all_dissections_top_weighted = defaultdict(float)
    sig_dissections_top_weighted = defaultdict(float)
    weighted_dissections_three_weighted = defaultdict(float)
    log_weighted_dissections_three_weighted = defaultdict(float)
    for row in sheet.iter_rows(min_row=2, min_col=18, max_col=18):
        for cell in row:
            data = cell.value 
            dissections_values = re.split(':|, ', data)
            p_value = sheet[f"G{cell.row }"].value
            # top one
            all_dissections_top_weighted[dissections_values[0]] += percent_to_float(dissections_values[1])
            if cell.row <= cuttoff+1:
                sig_dissections_top_weighted[dissections_values[0]] += percent_to_float(dissections_values[1])
            # top three
            for i in range (0, len(dissections_values), 2):
                # all dissections
                all_dissections_three_count[dissections_values[i]] += 1
                all_dissections_three_weighted[dissections_values[i]] += percent_to_float(dissections_values[i+1])
                # sig dissections
                if cell.row <= cuttoff+1:
                    sig_dissections_three_count[dissections_values[i]] += 1
                    sig_dissections_three_weighted[dissections_values[i]] += percent_to_float(dissections_values[i+1])
                # proability weighted for dissections
                weighted_dissections_three_weighted[dissections_values[i]] += percent_to_float(dissections_values[i+1])/p_value
                log_weighted_dissections_three_weighted[dissections_values[i]] -= percent_to_float(dissections_values[i+1]) * np.log(p_value)

    # float cleaning
    for key in all_region_three_weighted:
        all_region_three_weighted[key] = round(all_region_three_weighted[key],3)
    for key in sig_region_three_weighted:
        sig_region_three_weighted[key] = round(sig_region_three_weighted[key],3)
    for key in all_region_top_weighted:
        all_region_top_weighted[key] = round(all_region_top_weighted[key],3)
    for key in sig_region_three_weighted:
        sig_region_top_weighted[key] = round(sig_region_top_weighted[key],3)
    
    for key in all_dissections_three_weighted:
        all_dissections_three_weighted[key] = round(all_dissections_three_weighted[key],3)
    for key in sig_dissections_three_weighted:
        sig_dissections_three_weighted[key] = round(sig_dissections_three_weighted[key],3)
    for key in all_dissections_top_weighted:
        all_dissections_top_weighted[key] = round(all_dissections_top_weighted[key],3)
    for key in sig_dissections_three_weighted:
        sig_dissections_top_weighted[key] = round(sig_dissections_top_weighted[key],3)
    
    chi_square_sig_three_region = chi_square(all_region_three_weighted, sig_region_three_weighted, cleaning=True)
    chi_square_weighted_p_region = chi_square(all_region_three_weighted, weighted_region_three_weighted)
    chi_square_sig_three_dissections = chi_square(all_dissections_three_weighted, sig_dissections_three_weighted, cleaning=True)
    chi_square_weighted_p_dissections = chi_square(all_dissections_three_weighted, weighted_dissections_three_weighted)

    # order fixing
    all_region_three_count_list, sig_region_three_count_list = reorder_dict(dict(all_region_three_count), dict(sig_region_three_count))
    all_region_three_weighted_list, sig_region_three_weighted_list = reorder_dict(dict(all_region_three_weighted), dict(sig_region_three_weighted))
    all_region_top_weighted_list, sig_region_top_weighted_list = reorder_dict(dict(all_region_top_weighted), dict(sig_region_top_weighted))

    all_dissections_three_count_list, sig_dissections_three_count_list = reorder_dict(dict(all_dissections_three_count), dict(sig_dissections_three_count), debugFlag=True)
    all_dissections_three_weighted_list, sig_dissections_three_weighted_list = reorder_dict(dict(all_dissections_three_weighted), dict(sig_dissections_three_weighted), debugFlag=True)
    all_dissections_top_weighted_list, sig_dissections_top_weighted_list = reorder_dict(dict(all_dissections_top_weighted), dict(sig_dissections_top_weighted), debugFlag=True)

    name_dict_list = [
                        ("Top 3 Count for All Clusters (Regions)", all_region_three_count_list),
                        ("Top 3 Count for Sig Clusters (Regions)", sig_region_three_count_list),
                        ("Top 3 Weighted for All Clusters (Regions)", all_region_three_weighted_list,),
                        ("Top 3 Weighted for Sig Clusters (Regions)", sig_region_three_weighted_list),
                        ("Top 1 Weighted for All Clusters (Regions)", all_region_top_weighted_list),
                        ("Top 1 Weighted for Sig Clusters (Regions)", sig_region_top_weighted_list),
                        ("Top 3 Weighted for P-Value Weighted Clusters (Regions)", weighted_region_three_weighted.items()),
                        ("Top 3 Weighted for Log P-Value Weighted Clusters (Regions)", log_weighted_region_three_weighted.items()),
                        ("Top 3 Count for All Clusters (Dissections)", all_dissections_three_count_list),
                        ("Top 3 Count for Sig Clusters (Dissections)", sig_dissections_three_count_list),
                        ("Top 3 Weighted for All Clusters (Dissections)", all_dissections_three_weighted_list,),
                        ("Top 3 Weighted for Sig Clusters (Dissections)", sig_dissections_three_weighted_list),
                        ("Top 1 Weighted for All Clusters (Dissections)", all_dissections_top_weighted_list),
                        ("Top 1 Weighted for Sig Clusters (Dissections)", sig_dissections_top_weighted_list),
                        ("Top 3 Weighted for P-Value Weighted Clusters (Dissections)", weighted_dissections_three_weighted.items()),
                        ("Top 3 Weighted for Log P-Value Weighted Clusters (Dissections)", log_weighted_dissections_three_weighted.items())
                     ]
    
    if not edit:
        # insert custom testing code
        print(all_dissections_three_count, "\n")
        print(len(all_dissections_three_count))
        return

    with open(f'cell_breakdown_images/{disorder}-{significance}-cell_type_breakdown.txt', 'w') as f:
        f.write(f"There are {cuttoff} significant clusters at the alpha={significance} level! \n")
        for name_dict in name_dict_list:
            name = name_dict[0]
            dictionary = name_dict[1]
            f.write(f"{name}: {dictionary} \n")
        f.write(f"Really sketchy chi-square result for the distribution of top 3 region weights {chi_square_sig_three_region} \n")
        f.write(f"Really sketchy chi-square result for the distribution of p-value weighted region weights {chi_square_weighted_p_region} \n")
        f.write(f"Really sketchy chi-square result for the distribution of top 3 dissection weights {chi_square_sig_three_dissections} \n")
        f.write(f"Really sketchy chi-square result for the distribution of p-value weighted dissection weights {chi_square_weighted_p_dissections} \n")
    
    for name_dict in name_dict_list:
        name = name_dict[0]
        dictionary_list = name_dict[1]
        keys = list(map(lambda x: x[0], dictionary_list))
        values = list(map(lambda x: x[1], dictionary_list))
        fig = plt.figure(figsize = (30, 8))
        plt.bar(keys, values, color ='maroon', width = 0.4)
        plt.title(name)
        plt.xticks(rotation=30, ha='right')
        plt.savefig(f'cell_breakdown_images/{disorder}-{significance}-{name}.png')

    assert sum(all_region_three_count.values()) == 3 * 461
    # print(sum(all_dissections_three_count.values())) = 1380 
    # NOTE: Row 178 only has one value
    # assert sum(all_dissections_three_count.values()) == 3 * 461


if __name__ == "__main__":
    get_all_dissections (check_against_dataset = True)
    # get_cell_breakdown(1.1931E-06, edit = False)