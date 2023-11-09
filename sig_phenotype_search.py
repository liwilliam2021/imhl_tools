import openpyxl
from collections import defaultdict

ALPHA = 0.05

def list_union(lst1, lst2):
    return list(set(lst1) | set(lst2))

def search_all_phenotypes (top = 50, significance = -1):

    wb = openpyxl.load_workbook(filename="EDITED_SILETTI_MAGMA_GSA_12_18_22.xlsx")

    cluster_data = defaultdict(lambda: {"sig_num_pheno": 0, "sig_pheno_w_prob": [],
                                        "top_num_pheno": 0, "top_pheno_w_prob": []})
    num_cluster = wb.worksheets[1].max_row - 1
    bf_adjusted_alpha = ALPHA / num_cluster

    # print(num_cluster, bf_adjusted_alpha)

    if significance == -1:
        significance = bf_adjusted_alpha

    for sheet in wb:
        if sheet.title in ["Overview", "SCZ_2011", "SCZ_2014", "SCZ_2018", "HFV2_LESS_STRICT", "HFV2_BROAD"]:
            continue
        phenotype = sheet.title
        break_flag = False
        #find the cuttoff row for significance
        for row in sheet.iter_rows(min_row=2, min_col=7, max_col=7):
            for cell in row:
                cluster= sheet[f"J{cell.row }"].value
                break_flag = True
                if cell.value < significance:
                    cluster_data[cluster]["sig_num_pheno"] += 1
                    cluster_data[cluster]["sig_pheno_w_prob"].append({phenotype: cell.value})
                    break_flag = False
                if cell.row - 1 <= top:
                    cluster_data[cluster]["top_num_pheno"] += 1
                    cluster_data[cluster]["top_pheno_w_prob"].append({phenotype: cell.value})
                    break_flag = False
            if break_flag == True:
                break

    o_wb = openpyxl.Workbook()
    ws = o_wb.active
    ws.title = "ClustersPhenotypes"
    ws['A1'] = "cluster"
    ws['B1'] = "all_phenotype_hits"
    ws['C1'] = "number_phenotypes"
    ws['D1'] = "sig_phenotypes"
    ws['E1'] = "num_sig_phenotypes"
    ws['F1'] = f"top{top}_phenotypes"
    ws['G1'] = f"num_top{top}_phenotypes"

    i = 2
    for cluster in cluster_data:
        ws[f"A{i}"] = cluster
        ws[f"D{i}"] = str(cluster_data[cluster]["sig_pheno_w_prob"])
        ws[f"E{i}"] = cluster_data[cluster]["sig_num_pheno"]
        ws[f"F{i}"] = str(cluster_data[cluster]["top_pheno_w_prob"])
        ws[f"G{i}"] = cluster_data[cluster]["top_num_pheno"]
        sig_list = [list(item.keys())[0] for item in cluster_data[cluster]["sig_pheno_w_prob"]]
        top_list = [list(item.keys())[0] for item in cluster_data[cluster]["top_pheno_w_prob"]]
        total_list = list_union(sig_list, top_list)
        ws[f"B{i}"] = str(total_list)
        ws[f"C{i}"] = len(total_list)
        i += 1
    
    o_wb.save("most_implicated_across_pheno.xlsx")
    print ("done")


if __name__ == "__main__":
    search_all_phenotypes ()